from customtkinter import *
from PIL import Image
from pathlib import Path
from tkinter import Tk, Canvas, Entry, Text, Button, PhotoImage, Scrollbar,ttk

import openpyxl as xl

import google.generativeai as genai
from deep_translator import *
from openpyxl.reader.excel import load_workbook
from openpyxl.workbook import Workbook

API_KEY = 'AIzaSyAfiWyhFw_UbggIionRb-_5pz91NyfARTk'
genai.configure(
    api_key=API_KEY
)
model = genai.GenerativeModel('gemini-pro')
chat = model.start_chat(history=[])

root = CTk()
window2 = CTk()

root.attributes("-fullscreen", True)
root.config(bg='#003C43')

logo = CTkImage(Image.open("image/logo.png"), size=(634, 405))

thefont = CTkFont(size=72, family="Agency FB")
thefonta = CTkFont(size=38, family="Agency FB")
thefont_1 = CTkFont(size=65, family="Agency FB")

thefont1 = CTkFont(size=18, family="HP Simplified")
thefont11 = CTkFont(size=65, family="HP Simplified")

thefont2 = CTkFont(size=18, family="HP Simplified")
thefont12 = CTkFont(size=65, family="HP Simplified")

his_font = CTkFont(size=35, family="Agency FB")
history = CTkFont(size=55, family="Agency FB")

the_homefont = CTkFont(size=40,family="Segoe UI Black")

thefile = 'Logins.xlsx'
wb = xl.load_workbook(thefile)
sheet = wb['Sheet1']

def on_button_click():
    print("Button clicked")
def on_enter_key(event=None):
    on_button_click()  # Call the button click function
try:
    workbook = load_workbook('Logins.xlsx')
    sheet = workbook.active
except FileNotFoundError:
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(['Username', 'Password'])  # Add headers if file is new
def add_user_to_excel(username1, password1, confirm_password1):
    if password1 != confirm_password1:

        entry_32.delete(0, 'end')
        return entry_32

    usernames = [cell.value for cell in sheet['A']]

    if username1 in usernames:


        entry_12.delete(0, 'end')
        return entry_12

    sheet.append([username1, password1])
    workbook.save('Logins.xlsx')
    print(f"User {username1} added successfully.")
def signask():
    global username, password, confirmp
    username = str(entry_12.get())
    password = str(entry_22.get())
    confirmp = str(entry_32.get())
    print(username, password, confirmp)

    if password == confirmp:
        result = add_user_to_excel(username, password, confirmp)
        if result:  # If entry_12 was cleared
            return
        else:
            ask3()
    else:
        entry_32.delete(0, 'end')
        entry_22.delete(0, 'end')
        return entry_32, entry_22
def confirmuser(usernamee1, passwordd1, attempts=3):
    for row in range(1, sheet.max_row + 1):
        if str(sheet.cell(row=row, column=1).value) == usernamee1:
            for _ in range(attempts):
                if str(sheet.cell(row=row, column=2).value) == passwordd1:
                    return True  # Return True if password matches
                else:
                    entry_21.delete(0, 'end')  # Clear the password text box
                    return False  # Password doesn't match; return False
            # If we reach this point, the username matched but password didn't
            break  # Exit the loop if username matches
    else:
        entry_11.delete(0, 'end')  # Clear the username text box
        return False  # Username not found; return False
def logask():
    global username2, password2
    username2 = str(entry_11.get())
    password2 = str(entry_21.get())



    if confirmuser(username2, password2):

        ask2()  # Log in successful; proceed to next step
    else:
        entry_11.delete(0, 'end')  # Clear the username text box
        # Handle authentication failure (e.g., display an error message)
def ask3():


    # Close the workbook
    workbook.close()
    window_frame1.destroy()
    OUTPUT_PATH = Path(__file__).parent
    ASSETS_PATH = OUTPUT_PATH / Path("assets/frame0")
    def both1():
        update_history1()
        ai()

    def relative_to_assets(path: str) -> Path:
        return ASSETS_PATH / Path(path)


    thefont = CTkFont(size=18, family="HP Simplified")
    thefont_1 = CTkFont(size=35, family="HP Simplified")


    window = CTkFrame(root, fg_color="red")
    window.place(relx=0.5, rely=0.5, relwidth=1, relheight=1, anchor="center")

    canvas = Canvas(
        window,
        bg="#003C43",
        height=1080,
        width=1980,
        bd=0,
        highlightthickness=0,
        relief="ridge"
    )
    canvas.place(x=0, y=0)
    image_image_1 = PhotoImage(
        file=relative_to_assets("image_1.png"))
    global image_1
    image_1 = CTkLabel(
        root, text="", fg_color="#135D66",
        image=image_image_1
    )
    image_1.place(x=1197.9999999999995,
                  y=350.0, anchor="center")
    canvas.create_rectangle(
        0,
        0.0,
        523.0000000000005,
        1080.0,
        fill="#135D66",
        outline="")
    image_image_2 = PhotoImage(
        file=relative_to_assets("image_2.png"))
    global image_2
    image_2 = CTkLabel(
        root, text="", fg_color="#003C43",
        image=image_image_2
    )
    image_2.place(x=1197.9999999999995,
                  y=601.0, anchor="center")
    image_image_3 = PhotoImage(
        file=relative_to_assets("image_3.png"))
    image_3 = CTkLabel(
        root, text="", fg_color="#135D66",
        image=image_image_3
    )
    image_3.place(x=260.99999999999955,
                  y=111.0, anchor="center")

    entry_image_1 = PhotoImage(
        file=relative_to_assets("entry_1.png"))
    entry_bg_1 = CTkLabel(
        root, text="", fg_color="#003C43",
        image=entry_image_1
    )
    entry_bg_1.place(x=1198.4999999999995,
                     y=961.5, anchor="center")
    global entry_1
    entry_1 = Entry(
        bd=0,
        bg="#003C43",
        fg="#E3FEF7",
        highlightthickness=0,
        font=thefont_1,
    )
    entry_1.place(
        x=728.9999999999995,
        y=932.0,
        width=890.0,
        height=60.0
    )

    button_image_1 = PhotoImage(
        file=relative_to_assets("button_1.png"))
    button_1 = CTkButton(
        root,
        text="",
        image=button_image_1,
        command=both1,
        fg_color="#003C43",
        bg_color="#003C43",
        hover_color="#003C43"
    )
    root.bind('<Return>', lambda event: both1())
    button_1.place(
        x=1605.9999999999995,
        y=930.0,
        relwidth=0.034
    )

    history_text = Text(
        root,
        bd=0,
        bg="#135D66",
        fg="#77B0AA",
        highlightthickness=0,
        font=thefont_1,
    )
    history_text.place(x=250, y=250, width=260, height=300)
    lab = CTkLabel(root, text="History", fg_color='#135D66', font=history, text_color="#E3FEF7")
    lab.place(relx=0.05, rely=0.18)
    history_frame = CTkScrollableFrame(root, fg_color='#135D66', bg_color='#135D66',
                                       scrollbar_button_color='#0E434A',
                                       scrollbar_button_hover_color='#092C31')
    history_frame.place(relx=0, rely=0.25, relwidth=0.272, relheight=0.74)

    lang = CTkButton(root, text="A/OROMOO",
                     text_color="#77B0AA", corner_radius=5, bg_color="#003C43", hover_color="#135D66",
                     fg_color="#003C43", border_width=2,
                     border_color="#77B0AA", command=ai)
    lang.place(relx=0.9, rely=0.86, relheight=0.03)
    lang1 = CTkButton(root, text="ENGLISH",
                      text_color="#77B0AA", corner_radius=5, bg_color="#003C43", hover_color="#135D66",
                      fg_color="#003C43", border_width=2,
                      border_color="#77B0AA", command=ai1)
    lang1.place(relx=0.9, rely=0.897, relheight=0.03)

    home_button = CTkButton(root, text="HOME", font=the_homefont, text_color="#77B0AA", command=welcome,
                            fg_color="#003C43", bg_color="#003C43", hover_color="#003C43")
    home_button.place(relx=0.83, rely=0.06, relheight=0.04)

    exit_button = CTkButton(root, text="EXIT", command=root.quit, text_color="#77B0AA", font=the_homefont,
                            fg_color="#003C43", bg_color="#003C43", hover_color="#003C43")
    exit_button.place(relx=0.9, rely=0.06, relheight=0.04)

    def update_history1():
        global question
        # Get the current value from entry_1
        new_entry = entry_1.get()
        question = entry_1.get()
        print(question)
        if new_entry:


            # Clear entry_1 after appending
            entry_1.delete(0, 'end')
    # Function to update history_text with entry_1 values

        histo = question

        for row in range(1, sheet.max_row + 1):
            if str(sheet.cell(row=row, column=1).value) == username and str(
                    sheet.cell(row=row, column=2).value) == password:

                # Find the first empty cell in the row
                for col in range(3, sheet.max_column + 1):
                    if sheet.cell(row=row, column=col).value is None:
                        sheet.cell(row=row, column=col).value = histo
                        break
                else:
                    # If no empty cell was found, add to the next column
                    sheet.cell(row=row, column=sheet.max_column + 1).value = histo




            # Close the workbook
        workbook.close()

        # Display the history in the GUI
        history = PhotoImage(file=relative_to_assets("www.png"))
        # history_text = "\n".join(entry for entry in this_history) if this_history else ""
        # history_bg = CTkButton(
        #     root, text=history_text, fg_color="#135D66",
        #     image=history
        # )
        k=""
        # history_bg.place(x=250, y=210, anchor="center")
        for i in histo:
            k+=i
            # his_but =CTkButton(history_frame, text=f'{i}',anchor='w', fg_color='#11235A',
            #                                 hover_color='#4960A9', width=500, text_color='#C6CF9B',
            #                                 )
            # his_but.grid(row=row, column=1, pady=5)
        but = CTkButton(history_frame, text=k, anchor='w', fg_color='#77B0AA',
                            hover_color='#209BAC', width=500, height=45, text_color='#E3FEF7', font=his_font)
        but.pack(pady=4)

        print(f"Entry '{histo}' added successfully.")
        workbook.save('Logins.xlsx')
def ask2():

        window_frame1.destroy()
        OUTPUT_PATH = Path(__file__).parent
        ASSETS_PATH = OUTPUT_PATH / Path("assets/frame0")
        # Initialize a variable to store the history


        # Iterate through each row in the work
        # Close the workbook
        workbook.close()
        def both():
            ai()
            update_history2()
            for row in sheet.iter_rows(min_row=2, values_only=True):  # Assuming the data starts from the second row
                usernames = row[0]
                historys = row[2:]  # Exclude the first two columns (username and password)

                # Check if the username matches 'Eyuel'
                if usernames == username2:
                    # Filter out None values from history
                    this_history = [entry for entry in historys if entry is not None]
                    break  # Exit the loop once the user is found

            # Print or process the extracted history
            if this_history:
                # Print the history line by line
                print(f"History for user {username2}:")
                for entry in this_history:
                    print(entry)
            else:
                print("User not found or has no history")

            # Close the workbook
            workbook.close()

            # Display the history in the GUI
            history = PhotoImage(file=relative_to_assets("www.png"))
            # history_text = "\n".join(entry for entry in this_history) if this_history else ""
            # history_bg = CTkButton(
            #     root, text=history_text, fg_color="#135D66",
            #     image=history
            # )

            #history_bg.place(x=250, y=210, anchor="center")
            for i in this_history:
                # his_but =CTkButton(history_frame, text=f'{i}',anchor='w', fg_color='#11235A',
                #                                 hover_color='#4960A9', width=500, text_color='#C6CF9B',
                #                                 )
                # his_but.grid(row=row, column=1, pady=5)
                but = CTkButton(history_frame,text=i,anchor='w', fg_color='#77B0AA',
                               hover_color='#209BAC', width=500,height=45, text_color='#E3FEF7',font=his_font)
                but.pack(pady=4)

        def relative_to_assets(path: str) -> Path:
            return ASSETS_PATH / Path(path)

        thefont = CTkFont(size=18, family="HP Simplified")
        thefont_1 = CTkFont(size=35, family="HP Simplified")



        window = CTkFrame(root, fg_color="red")
        window.place(relx=0.5, rely=0.5, relwidth=1, relheight=1, anchor="center")

        canvas = Canvas(
            window,
            bg="#003C43",
            height=1080,
            width=1980,
            bd=0,
            highlightthickness=0,
            relief="ridge"
        )
        canvas.place(x=0, y=0)
        image_image_1 = PhotoImage(
            file=relative_to_assets("image_1.png"))
        global image_1
        image_1 = CTkLabel(
            root, text="", fg_color="#135D66",
            image=image_image_1
        )
        image_1.place(x=1197.9999999999995,
                      y=350.0, anchor="center")
        canvas.create_rectangle(
            0,
            0.0,
            523.0000000000005,
            1080.0,
            fill="#135D66",
            outline="")

        image_image_2 = PhotoImage(
            file=relative_to_assets("image_2.png"))
        global image_2
        image_2 = CTkLabel(
            root, text="", fg_color="#003C43",
            image=image_image_2
        )
        image_2.place(x=1197.9999999999995,
                      y=601.0, anchor="center")
        image_image_3 = PhotoImage(
            file=relative_to_assets("image_3.png"))
        image_3 = CTkLabel(
            root, text="", fg_color="#135D66",
            image=image_image_3
        )
        image_3.place(x=260.99999999999955,
                      y=111.0, anchor="center")

        entry_image_1 = PhotoImage(
            file=relative_to_assets("entry_1.png"))
        entry_bg_1 = CTkLabel(
            root, text="", fg_color="#003C43",
            image=entry_image_1
        )
        entry_bg_1.place(x=1198.4999999999995,
                         y=961.5, anchor="center")
        global entry_1
        entry_1 = Entry(
            bd=0,
            bg="#003C43",
            fg="#E3FEF7",
            highlightthickness=0,
            font=thefont_1,
        )
        entry_1.place(
            x=728.9999999999995,
            y=932.0,
            width=890.0,
            height=60.0
        )

        button_image_1 = PhotoImage(
            file=relative_to_assets("button_1.png"))
        button_1 = CTkButton(
            root,
            text="",
            image=button_image_1,
            command=both,
            fg_color="#003C43",
            bg_color="#003C43",
            hover_color="#003C43"
        )
        root.bind('<Return>', lambda event: both())
        button_1.place(
            x=1605.9999999999995,
            y=930.0,
            relwidth=0.034
        )
        lab = CTkLabel(root,text="History",fg_color='#135D66',font=history,text_color="#E3FEF7")
        lab.place(relx=0.05, rely=0.18)
        history_frame =CTkScrollableFrame(root, fg_color='#135D66',bg_color='#135D66', scrollbar_button_color='#0E434A',
                                                scrollbar_button_hover_color='#092C31')
        history_frame.place(relx=0, rely=0.25, relwidth=0.272, relheight=0.74)

        lang = CTkButton(root,text="A/OROMOO",
                         text_color="#77B0AA",corner_radius=5,bg_color="#003C43",hover_color="#135D66",
                         fg_color="#003C43",border_width=2,
                         border_color="#77B0AA",command=ai)
        lang.place(relx=0.9,rely=0.86,relheight=0.03)
        lang1 = CTkButton(root, text="ENGLISH",
                          text_color="#77B0AA",corner_radius=5,bg_color="#003C43",hover_color="#135D66",
                          fg_color="#003C43",border_width=2,
                          border_color="#77B0AA",command=ai1)
        lang1.place(relx=0.9, rely=0.897,relheight=0.03)


        home_button = CTkButton(root, text="HOME",font=the_homefont,text_color="#77B0AA",command=welcome,fg_color="#003C43",bg_color="#003C43",hover_color="#003C43")
        home_button.place(relx=0.83, rely=0.06,relheight=0.04)

        exit_button = CTkButton(root, text="EXIT", command=root.quit,text_color="#77B0AA",font=the_homefont,fg_color="#003C43",bg_color="#003C43",hover_color="#003C43")
        exit_button.place(relx=0.9, rely=0.06,relheight=0.04)

        def update_history2():
            global question
            # Get the current value from entry_1
            new_entry = entry_1.get()
            question = entry_1.get()
            print(question)
            if new_entry:
                # Clear entry_1 after appending
                entry_1.delete(0, 'end')
            # Function to update history_text with entry_1 values

            histo = question

            for row in range(1, sheet.max_row + 1):
                if str(sheet.cell(row=row, column=1).value) == username2 and str(
                        sheet.cell(row=row, column=2).value) == password2:

                    # Find the first empty cell in the row
                    for col in range(3, sheet.max_column + 1):
                        if sheet.cell(row=row, column=col).value is None:
                            sheet.cell(row=row, column=col).value = histo
                            break
                    else:
                        # If no empty cell was found, add to the next column
                        sheet.cell(row=row, column=sheet.max_column + 1).value = histo

            print(f"Entry '{histo}' added successfully.")
            workbook.save('Logins.xlsx')
def ai1():

    image_1.destroy()
    image_2.destroy()

    # Schedule the rest of the ai function to be executed after a short delay

    def print_coordinates(event):

        # Bind the function to the Motion event
        root.bind("<Motion>", print_coordinates)

    root.after(1000, rest_of_ai1)
def rest_of_ai1():
    global frame3

    frame3 = CTkFrame(root, fg_color="#003C43", border_color="#003C43")  # Use a light grey background
    frame3.place(relwidth=0.517, relheight=0.64, x=693, y=130)

    hi = GoogleTranslator(source='auto', target="en").translate(question)
    if True:
        response = chat.send_message(hi)
        print('\n')
        ao = "en"
        tr = GoogleTranslator(source='en', target=ao).translate(response.text)
        thegeneratedtext = ""
        for i in tr:
            if i == '*':
                i = i.replace('*', " ")
            thegeneratedtext = thegeneratedtext + i
        tr = thegeneratedtext
        # Create a Text widget for the chat window
        text = Text(frame3, state='disabled', wrap='word', bg='#003C43', fg='#E3FEF7',
                    font=thefonta, bd=0)  # Use a larger font
        text.place(relx=0.5, rely=0.5, relwidth=1, relheight=1, anchor="center")

        # Create a Scrollbar and pack it on the right side of the Text widget
        scrollbar = CTkScrollbar(frame3, command=text.yview,
                                 fg_color='#003C43')  # Use a light grey background for the scrollbar
        scrollbar.pack(side='right', fill='y')

        # Configure the Text widget to use the Scrollbar
        text.config(yscrollcommand=scrollbar.set)

        # Load the image and keep a reference to it
        image_image_6 = PhotoImage(file="User.png")
        image_image_5 = PhotoImage(file="Odagp.png")
        text.image_reference1 = image_image_6
        text.image_reference2 = image_image_5

        # Insert the user's input and the translated response into the Text widget
        text.configure(state='normal')
        text.image_create('end', image=image_image_6)  # Insert image at the end
        text.insert('end',
                    "     " + "You " + "\n" + "                " + question + "\n" + "\n")  # Add space before "User:"
        text.image_create('end', image=image_image_5)
        text.insert('end', "\n" + "                " + tr + "\n")
        text.configure(state='disabled')


        print(tr)
        print('\n')
def ai():

    image_1.destroy()
    image_2.destroy()

    # Schedule the rest of the ai function to be executed after a short delay



    root.after(1000, rest_of_ai)
def rest_of_ai():
    global frame3

    frame3 = CTkFrame(root, fg_color="#003C43", border_color="#003C43")  # Use a light grey background
    frame3.place(relwidth=0.517, relheight=0.64, x=693, y=130)

    hi = GoogleTranslator(source='auto', target="en").translate(question)
    if True:
        response = chat.send_message(hi)
        print('\n')
        ao = "om"
        tr = GoogleTranslator(source='en', target=ao).translate(response.text)
        thegeneratedtext = ""
        for i in tr:
            if i == '*':
                i = i.replace('*', " ")
            thegeneratedtext = thegeneratedtext + i
        tr = thegeneratedtext
        # Create a Text widget for the chat window
        text = Text(frame3, state='disabled', wrap='word', bg='#003C43', fg='#E3FEF7',
                    font=thefonta, bd=0)  # Use a larger font
        text.place(relx=0.5, rely=0.5, relwidth=1, relheight=1, anchor="center")

        # Create a Scrollbar and pack it on the right side of the Text widget
        scrollbar = CTkScrollbar(frame3, command=text.yview,
                                 fg_color='#003C43')  # Use a light grey background for the scrollbar
        scrollbar.pack(side='right', fill='y')

        # Configure the Text widget to use the Scrollbar
        text.config(yscrollcommand=scrollbar.set)

        # Load the image and keep a reference to it
        image_image_6 = PhotoImage(file="User.png")
        image_image_5 = PhotoImage(file="Odagp.png")
        text.image_reference1 = image_image_6
        text.image_reference2 = image_image_5

        # Insert the user's input and the translated response into the Text widget
        text.configure(state='normal')
        text.image_create('end', image=image_image_6)  # Insert image at the end
        text.insert('end',
                    "     " + "You " + "\n" + "                " + question + "\n" + "\n")  # Add space before "User:"
        text.image_create('end', image=image_image_5)
        text.insert('end', "\n" + "                " + tr + "\n")
        text.configure(state='disabled')


        print(tr)
        print('\n')
def sign():
    window_frame1.destroy()

    OUTPUT_PATH = Path(__file__).parent
    ASSETS_PATH = OUTPUT_PATH / Path("assets2/frame0")

    def relative_to_assets(path: str) -> Path:
        return ASSETS_PATH / Path(path)

    global window_frame2
    window_frame2 = CTkFrame(root, fg_color="red")
    window_frame2.place(relx=0.5, rely=0.5, relwidth=1, relheight=1, anchor="center")

    canvas2 = Canvas(
        window_frame2,
        bg="#003C43",
        height=1080,
        width=1980,
        bd=0,
        highlightthickness=0,
        relief="ridge"
    )

    canvas2.place(x=0, y=0)
    canvas2.create_rectangle(
        1113.0,
        0.0,
        1980.0,
        1080.0,
        fill="#135D66",
        outline="")

    image_image_12 = PhotoImage(
        file=relative_to_assets("image_1.png"))
    image_12 = CTkLabel(
        root, text="", fg_color="#135D66",
        image=image_image_12
    )
    image_12.place(x=1312.0,
                   y=288.0, anchor="center")
    entry_image_12 = PhotoImage(
        file=relative_to_assets("entry_1.png"))
    entry_bg_12 = CTkLabel(
        root, text="", fg_color="#135D66",
        image=entry_image_12
    )
    entry_bg_12.place(x=1486.0,
                      y=374.0, anchor="center")
    global entry_12
    entry_12 = Entry(
        bd=0,
        bg="#E3FEF7",
        fg="#000716",
        highlightthickness=0,
        foreground="#77B0AA",
        font=thefont12,
    )
    entry_12.place(
        x=1243.0,
        y=320.0,
        width=486.0,
        height=106.0
    )

    entry_image_22 = PhotoImage(
        file=relative_to_assets("entry_2.png"))
    entry_bg_22 = CTkLabel(
        root, text="", fg_color="#135D66",
        image=entry_image_22
    )
    entry_bg_22.place(x=1486.0,
                      y=548.0, anchor="center")
    global entry_22
    entry_22 = Entry(
        bd=0,
        bg="#E3FEF7",
        fg="#000716",
        highlightthickness=0,
        foreground="#77B0AA",
        font=thefont12,
        show="*"
    )
    entry_22.place(
        x=1243.0,
        y=494.0,
        width=486.0,
        height=106.0
    )

    entry_image_32 = PhotoImage(
        file=relative_to_assets("entry_3.png"))
    entry_bg_32 = CTkLabel(
        root, text="", fg_color="#135D66",
        image=entry_image_32
    )
    entry_bg_32.place(x=1486.0,
                      y=722.0, anchor="center")
    global entry_32
    entry_32 = Entry(
        bd=0,
        bg="#E3FEF7",
        fg="#000716",
        highlightthickness=0,
        foreground="#77B0AA",
        font=thefont12,
        show="*"
    )
    entry_32.place(
        x=1243.0,
        y=668.0,
        width=486.0,
        height=106.0
    )

    image_image_22 = PhotoImage(
        file=relative_to_assets("image_2.png"))
    image_22 = CTkLabel(
        root, text="", fg_color="#135D66",
        image=image_image_22
    )
    image_22.place(x=1546.0,
                   y=140.0, anchor="center")
    button_12 = CTkButton(
        window_frame2,
        text="Sign Up",
        font=thefont12,
        command=signask,
        bg_color="#135D66",
        fg_color="#448489",
        hover_color="#5FB3BA"
    )
    root.bind('<Return>', lambda event: signask())
    button_12.place(
        x=1234.0,
        y=880.0,
        relwidth=0.263,
        relheight=0.1
    )

    image_image_32 = PhotoImage(
        file=relative_to_assets("image_3.png"))
    image_32 = CTkLabel(
        root, text="", fg_color="#135D66",
        image=image_image_32
    )
    image_32.place(x=1243.0,
                   y=605.0)

    image_image_42 = PhotoImage(
        file=relative_to_assets("image_4.png"))
    image_42 = CTkLabel(
        root, text="", fg_color="#135D66",
        image=image_image_42
    )
    image_42.place(x=1243.0,
                   y=432.0)
    image_image_52 = PhotoImage(
        file=relative_to_assets("image_5.png"))
    image_52 = CTkLabel(
        root, text="", fg_color="#003C43",
        image=image_image_52
    )
    image_52.place(x=590.0,
                   y=523.0, anchor="center")
    button_22 = Button(
        # image=button_image_2,
        bg="#135D66",
        text="Login",
        underline=6,
        fg="white",
        font=thefont2,
        borderwidth=0,
        highlightthickness=0,
        command=log1,
        relief="flat"
    )
    button_22.place(
        x=1630.0,
        y=188.0,
        width=75.0,
        height=30.0
    )


    def toggle_password_visibility(hide_image=None, show_image=None):
        global show_button
        show_image = PhotoImage(file="showp.png")
        hide_image = PhotoImage(file="hide.png")

        if entry_22.cget('show') == '':
            entry_22.config(show='*')
            show_button.config(image=show_image, compound="left")
        else:
            entry_22.config(show='')
            show_button.config(image=hide_image, compound="left")


        # Define the show_button globally

        show_button = Button(
        command=toggle_password_visibility,
        font=("Helvetica", 12)
    )
        show_button.place(
        x=1440.0,  # Adjust x and y to position the button appropriately
        y=494.0,
        width=512,
        height=512
    )

    # Call toggle_password_visibility to set the initial state
        toggle_password_visibility()
def log1():
    root_frame.destroy()
    window_frame2.destroy()

    OUTPUT_PATH1 = Path(__file__).parent
    ASSETS_PATH1 = OUTPUT_PATH1 / Path("assets1/frame0")

    def relative_to_assets(path: str) -> Path:
        return ASSETS_PATH1 / Path(path)

    global window_frame1
    window_frame1 = CTkFrame(root, fg_color="#003C43")
    window_frame1.place(relx=0.5, rely=0.5, relwidth=1, relheight=1, anchor="center")

    canvas1 = Canvas(
        root,
        bg="#003C43",
        height=1080,
        width=1980,
        bd=0,
        highlightthickness=0,
        relief="ridge"
    )

    canvas1.place(x=0, y=0)
    canvas1.create_rectangle(
        0.0,
        0.0,
        867.0,
        1080.0,
        fill="#135D66",
        outline="")
    # canvas1 = CTkFrame(window_frame1,
    #                    fg_color="red"
    #
    # )
    # canvas1.place(relwidth=0.4,relhieght=0.4,x=0,y=0)

    image_image_11 = PhotoImage(
        file=relative_to_assets("image_1.png"))
    image_11 = CTkLabel(
        root, text=" ", fg_color="#135D66",
        image=image_image_11
    )
    image_11.place(x=403.0,
                   y=138.0, anchor="center")
    image_image_21 = PhotoImage(
        file=relative_to_assets("image_2.png"))
    image_21 = CTkLabel(
        root, text=" ", fg_color="#135D66",
        image=image_image_21
    )
    image_21.place(x=428.0,
                   y=215.0, anchor="center")

    image_image_31 = PhotoImage(
        file=relative_to_assets("image_3.png"))
    image_31 = CTkLabel(
        root, text=" ", fg_color="#135D66",
        image=image_image_31
    )
    image_31.place(x=196.0,
                   y=388.0, anchor="center")

    image_image_41 = PhotoImage(
        file=relative_to_assets("image_4.png"))
    image_41 = CTkLabel(
        root, text=" ", fg_color="#135D66",
        image=image_image_41
    )
    image_41.place(x=257.0,
                   y=591.0, anchor="center")

    button_11 = CTkButton(
        root,
        text="Login",
        font=thefont11,
        command=logask,
        bg_color="#135D66",
        fg_color="#448489",
        hover_color="#5FB3BA"
    )

    button_11.place(
        x=118.0,
        y=823.0,
        relwidth=0.263,
        relheight=0.1
    )
    root.bind('<Return>', lambda event: logask())
    image_image_51 = PhotoImage(
        file=relative_to_assets("image_5.png"))
    image_51 = CTkLabel(
        root, text=" ", fg_color='#003C43',
        image=image_image_51
    )
    image_51.place(x=1430.0,
                   y=523.0, anchor="center")

    entry_image_11 = PhotoImage(
        file=relative_to_assets("entry_1.png"))
    entry_bg_11 = CTkLabel(
        root, text=" ", fg_color="#135D66",
        image=entry_image_11
    )
    entry_bg_11.place(x=370.0,
                      y=470.0, anchor="center")
    global entry_11
    entry_11 = Entry(
        bd=0,
        bg="#E3FEF7",
        fg="#000716",
        highlightthickness=0,
        foreground="#77B0AA",
        font=thefont11,
    )
    entry_11.place(
        x=127.0,
        y=416.0,
        width=486.0,
        height=106.0
    )

    entry_image_21 = PhotoImage(
        file=relative_to_assets("entry_2.png"))
    entry_bg_21 = CTkLabel(
        root, text=" ", fg_color="#135D66",
        image=entry_image_21
    )
    entry_bg_21.place(x=370.0,
                      y=678.0, anchor="center")
    global entry_21

    entry_21 = Entry(
        bd=0,
        bg="#E3FEF7",
        fg="#000716",
        highlightthickness=0,
        foreground="#77B0AA",
        font=thefont11,
        show="*"
    )

    entry_21.place(
        x=127.0,
        y=624.0,
        width=486.0,
        height=106.0
    )

    # button_image_21 = PhotoImage(
    #     \file=relative_to_assets("button_2.png"))
    button_21 = Button(
        root,
        # image=button_image_2,
        bg="#135D66",
        text="SignUp",
        underline=6,
        fg="white",
        font=thefont1,
        borderwidth=0,
        command=sign,
    )
    button_21.place(
        x=510.0,
        y=201.0,
        width=75.0,
        height=30.0
    )
def log():
    root_frame.destroy()
    # window_frame2.destroy()

    OUTPUT_PATH1 = Path(__file__).parent
    ASSETS_PATH1 = OUTPUT_PATH1 / Path("assets1/frame0")

    def relative_to_assets(path: str) -> Path:
        return ASSETS_PATH1 / Path(path)

    global window_frame1
    window_frame1 = CTkFrame(root, fg_color="#003C43")
    window_frame1.place(relx=0.5, rely=0.5, relwidth=1, relheight=1, anchor="center")

    canvas1 = Canvas(
        root,
        bg="#003C43",
        height=1080,
        width=1980,
        bd=0,
        highlightthickness=0,
        relief="ridge"
    )

    canvas1.place(x=0, y=0)
    canvas1.create_rectangle(
        0.0,
        0.0,
        867.0,
        1080.0,
        fill="#135D66",
        outline="")
    # canvas1 = CTkFrame(window_frame1,
    #                    fg_color="red"
    #
    # )
    # canvas1.place(relwidth=0.4,relhieght=0.4,x=0,y=0)

    image_image_11 = PhotoImage(
        file=relative_to_assets("image_1.png"))
    image_11 = CTkLabel(
        root, text=" ", fg_color="#135D66",
        image=image_image_11
    )
    image_11.place(x=403.0,
                   y=138.0, anchor="center")
    image_image_21 = PhotoImage(
        file=relative_to_assets("image_2.png"))
    image_21 = CTkLabel(
        root, text=" ", fg_color="#135D66",
        image=image_image_21
    )
    image_21.place(x=428.0,
                   y=215.0, anchor="center")

    image_image_31 = PhotoImage(
        file=relative_to_assets("image_3.png"))
    image_31 = CTkLabel(
        root, text=" ", fg_color="#135D66",
        image=image_image_31
    )
    image_31.place(x=196.0,
                   y=388.0, anchor="center")

    image_image_41 = PhotoImage(
        file=relative_to_assets("image_4.png"))
    image_41 = CTkLabel(
        root, text=" ", fg_color="#135D66",
        image=image_image_41
    )
    image_41.place(x=257.0,
                   y=591.0, anchor="center")

    button_11 = CTkButton(
        root,
        text="Login",
        font=thefont11,
        command=logask,
        bg_color="#135D66",
        fg_color="#448489",
        hover_color="#5FB3BA"
    )

    button_11.place(
        x=118.0,
        y=823.0,
        relwidth=0.263,
        relheight=0.1
    )
    root.bind('<Return>', lambda event: logask())
    image_image_51 = PhotoImage(
        file=relative_to_assets("image_5.png"))
    image_51 = CTkLabel(
        root, text=" ", fg_color='#003C43',
        image=image_image_51
    )
    image_51.place(x=1430.0,
                   y=523.0, anchor="center")

    entry_image_11 = PhotoImage(
        file=relative_to_assets("entry_1.png"))

    entry_bg_11 = CTkLabel(
        root, text=" ", fg_color="#135D66",
        image=entry_image_11
    )
    entry_bg_11.place(x=370.0,
                      y=470.0, anchor="center")
    global entry_11
    entry_11 = Entry(
        bd=0,
        bg="#E3FEF7",
        fg="#000716",
        highlightthickness=0,
        foreground="#77B0AA",
        font=thefont11,
    )
    entry_11.place(
        x=127.0,
        y=416.0,
        width=486.0,
        height=106.0
    )

    entry_image_21 = PhotoImage(
        file=relative_to_assets("entry_2.png"))
    entry_bg_21 = CTkLabel(
        root, text=" ", fg_color="#135D66",
        image=entry_image_21
    )
    entry_bg_21.place(x=370.0,
                      y=678.0, anchor="center")
    global entry_21
    entry_21 = Entry(
        bd=0,
        bg="#E3FEF7",
        fg="#000716",
        highlightthickness=0,
        foreground="#77B0AA",
        font=thefont11,
        show="*"
    )
    entry_21.place(
        x=127.0,
        y=624.0,
        width=486.0,
        height=106.0
    )

    # button_image_21 = PhotoImage(
    #     file=relative_to_assets("button_2.png"))
    button_21 = Button(
        root,
        # image=button_image_2,
        bg="#135D66",
        text="SignUp",
        underline=6,
        fg="white",
        font=thefont1,
        borderwidth=0,
        command=sign,
    )
    button_21.place(
        x=510.0,
        y=201.0,
        width=75.0,
        height=30.0
    )
def welcome():
    window_frame1.destroy()

    global root_frame
    root_frame = CTkFrame(root)
    root_frame.place(relx=0.5, rely=0.5, relwidth=1, relheight=1, anchor="center")
    frame1 = CTkFrame(root_frame, fg_color="#003C43")
    frame1.place(relx=0.5, rely=0.5, relwidth=1, relheight=1, anchor="center")
    logo_label = CTkLabel(root_frame, text="", image=logo, bg_color='#003C43')
    logo_label.place(x=308, y=319, relwidth=0.401010101010101, relheight=0.4101851851851852)

    line = CTkFrame(root_frame, bg_color="#77B0AA", fg_color="#77B0AA")
    line.place(x=1243, y=65, relwidth=0.0030303030303030303, relheight=0.8768518518518519)

    frame = CTkFrame(root_frame, bg_color="#135D66", fg_color="#135D66")
    frame.place(x=1288, y=0, relwidth=0.4191919191919192, relheight=1.0)

    label = CTkLabel(root_frame, text="Your 24/7", font=thefont, fg_color="#135D66", anchor="w")
    label.place(x=1397, y=229, relheight=0.07592592592592592)

    label1 = CTkLabel(root_frame, text="guide", font=thefont, fg_color="#135D66", anchor="w")
    label1.place(x=1397, y=327, relheight=0.07592592592592592)

    label2 = CTkLabel(root_frame, text="to", font=thefont, fg_color="#135D66", anchor="w")
    label2.place(x=1397, y=425, relheight=0.07592592592592592)

    label3 = CTkLabel(root_frame, text="seamless", font=thefont, fg_color="#135D66", anchor="w")
    label3.place(x=1397, y=523, relheight=0.07592592592592592)

    label = CTkLabel(root_frame, text="conversation!", font=thefont, fg_color="#135D66", anchor="w")
    label.place(x=1397, y=621, relheight=0.07592592592592592)

    continue_button = CTkButton(frame, text="CONTINUE",
                                font=thefont_1, anchor="center",
                                text_color="#77B0AA",
                                command=log,
                                height=122,
                                width=333,
                                corner_radius=1000,
                                fg_color="#003C43",
                                border_width=5,
                                border_color="#77B0AA",
                                hover_color="#005D68")
    continue_button.place(relx=0.4, rely=0.83, anchor="center")


global root_frame
root_frame = CTkFrame(root, fg_color='#003C43')
root_frame.place(relx=0.5, rely=0.5, relwidth=1, relheight=1, anchor="center")
logo_label = CTkLabel(root_frame, text="", image=logo, bg_color='#003C43')
logo_label.place(x=308, y=319, relwidth=0.401010101010101, relheight=0.4101851851851852)

line = CTkFrame(root_frame, bg_color="#77B0AA", fg_color="#77B0AA")
line.place(x=1243, y=65, relwidth=0.0030303030303030303, relheight=0.8768518518518519)

frame = CTkFrame(root_frame, bg_color="#135D66", fg_color="#135D66")
frame.place(x=1288, y=0, relwidth=0.4191919191919192, relheight=1.0)

label = CTkLabel(root_frame, text="Your 24/7", font=thefont, fg_color="#135D66", anchor="w")
label.place(x=1397, y=229, relheight=0.07592592592592592)

label1 = CTkLabel(root_frame, text="guide", font=thefont, fg_color="#135D66", anchor="w")
label1.place(x=1397, y=327, relheight=0.07592592592592592)

label2 = CTkLabel(root_frame, text="to", font=thefont, fg_color="#135D66", anchor="w")
label2.place(x=1397, y=425, relheight=0.07592592592592592)

label3 = CTkLabel(root_frame, text="seamless", font=thefont, fg_color="#135D66", anchor="w")
label3.place(x=1397, y=523, relheight=0.07592592592592592)

label = CTkLabel(root_frame, text="conversation!", font=thefont, fg_color="#135D66", anchor="w")
label.place(x=1397, y=621, relheight=0.07592592592592592)

continue_button = CTkButton(frame, text="CONTINUE",
                            font=thefont_1, anchor="center",
                            text_color="#77B0AA",
                            command=log,
                            height=122,
                            width=333,
                            corner_radius=1000,
                            fg_color="#003C43",
                            border_width=5,
                            border_color="#77B0AA",
                            hover_color="#005D68")
continue_button.place(relx=0.4, rely=0.83, anchor="center")

root.mainloop()